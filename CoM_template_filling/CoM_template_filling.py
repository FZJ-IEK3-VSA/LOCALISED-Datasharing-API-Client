import os
import sys

# Add parent directory to Python path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)


import re
import shutil
import time
from openpyxl import load_workbook, Workbook
import pandas as pd
import logging
import json

from zoomin_client import client


# Configure logger
logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("CoM_template_filling.log"),
        logging.StreamHandler(),  # Also print to console
    ],
)


def get_region_data(region_code, pathway_description="national", result_format="df"):
    """
    Get region data from the DSP
    """
    region_data = client.get_region_data(
        version="v4",
        country_code=region_code[:2].lower(),
        region_code=region_code,
        pathway_description=pathway_description,
        result_format=result_format,
    )
    return region_data


# Probability of hazard -  Low; Moderate; High; Not known
# Impact of hazard  -  Low; Moderate; High; Not known
# Expected change in hazard intensity - Increase; Decrease; No change; Not known
# Expected change in hazard frequency - Increase; Decrease; No change; Not known
# Timeframe(s) - Short-term; Mid-term; Long-term; Not known

probability_impact_string_mapping = {
    2: "High",
    1: "Moderate",
    0: "Low",
    -5: "Uncertain",
    -10: "Not known",
}

intensity_frequency_string_mapping = {
    1: "Increase",
    0: "No change",
    -1: "Decrease",
    -5: "Uncertain",
    -10: "Not known",
}

timeframe_string_mapping = {
    0: "Short-term",
    1: "Mid-term",
    2: "Long-term",
    -5: "Uncertain",
    -10: "Not known",
}


def extract_variables(equation):
    """
    Extract variables from an equation
    """
    # Define a regex pattern for variable names (letters, numbers, and underscores)
    variable_pattern = r"\b[a-zA-Z_][a-zA-Z0-9_]*\b"
    # Find all matches
    matches = re.findall(variable_pattern, equation)
    # Remove keywords that are not variables (e.g., "100" or operators)
    return [var.strip() for var in matches if not var.strip().isdigit()]


def get_dsp_value(variable, region_data, year=2020, climate_experiment="RCP8.5"):
    """
    Get value from DSP for a variable
    """
    try:
        if variable.startswith("eucalc_"):
            sub_region_data = region_data[
                (region_data["var_name"] == variable) & (region_data["year"] == year)
            ]  # 2020 value is taken

            if len(sub_region_data) == 0:
                logger.warning(f"No data found for variable {variable}")
                return None

            dsp_value = sub_region_data["value"].iloc[0]

        elif variable.startswith("cproj_"):
            sub_region_data = region_data[
                (region_data["var_name"] == variable)
                & (region_data["year"] == year)
                & (
                    region_data["climate_experiment"] == climate_experiment
                )  # 2020 value is taken
            ]  #  RCP4.5 scenario is considered

            if len(sub_region_data) == 0:
                logger.warning(f"No data found for variable {variable}")
                return None

            dsp_value = sub_region_data["value"].iloc[0]

        elif variable.startswith("cimp_"):
            sub_region_data = region_data[
                (region_data["var_name"] == variable)
                & (region_data["climate_experiment"].isin(("RCP4.5", "Historical")))
            ]  # Historical or RCP4.5 value is taken
            # for climate impact data

            if len(sub_region_data) == 0:
                logger.warning(f"No data found for variable {variable}")
                return None

            if ("historical_probability" in variable) or ("impact" in variable):
                int_value = int(sub_region_data["value"].iloc[0])
                dsp_value = probability_impact_string_mapping[int_value]

            elif ("change_in_frequency" in variable) or (
                "change_in_intensity" in variable
            ):
                int_value = int(sub_region_data["value"].iloc[0])
                dsp_value = intensity_frequency_string_mapping[int_value]

            elif "time_frame" in variable:
                int_value = int(sub_region_data["value"].iloc[0])
                dsp_value = timeframe_string_mapping[int_value]

        else:
            sub_region_data = region_data[region_data["var_name"] == variable]
            if len(sub_region_data) == 0:
                logger.warning(f"No data found for variable {variable}")
                return None

            dsp_value = sub_region_data["value"].iloc[0]

        return dsp_value

    except Exception as e:
        logger.error(f"Error getting DSP value for {variable}: {str(e)}")
        return 0


def calculate_sois(region_code: str, region_data: pd.DataFrame) -> dict:
    """
    Calculate SOIs for a region.
    """
    soi_df = pd.DataFrame(
        columns=[
            "soi_name",
            "var_name",
            "soi_description",
            "methodology",
            "SECAP_link",
            "SDG_targets",
            "var_unit",
            "value",
        ]
    )

    # Get SOI calculation excel sheet
    soi_metadata_df = pd.read_excel(
        os.path.join(
            current_dir,
            "data",
            "input",
            "variables_with_details_and_tags.xlsx",
        ),
        sheet_name="admin_business_and_social_KPIs",
    )

    # SOI calculations using collected and EUCalc data
    soi_vars_with_dsp_input = soi_metadata_df[
        (~soi_metadata_df["var_name"].isna())
        & (~soi_metadata_df["calculation"].isin(["BLANK", "TBD"]))
        & (  # some SOIs are decided to be left blank or TBD yet
            soi_metadata_df["data_source"] != "TOTAL"
        )  # some SOIs are sum of other SOIs
    ][
        [
            "soi_name",
            "var_name",
            "soi_description",
            "methodology",
            "SECAP_link",
            "SDG_targets",
            "var_unit",
            "calculation",
        ]
    ]

    for idx, row in soi_vars_with_dsp_input.iterrows():
        soi_name = row["soi_name"]
        soi_var_name = row["var_name"]
        soi_var_description = row["soi_description"]
        methodology = row["methodology"]
        SECAP_link = row["SECAP_link"]
        SDG_targets = row["SDG_targets"]
        var_unit = row["var_unit"]
        equation = row["calculation"]

        equation = equation.replace("\n", " ")

        try:
            # cases where calculations are required
            if any(symbol in equation for symbol in ["+", "/", "*"]):
                input_vars = extract_variables(equation)

                for input_var in input_vars:
                    value = get_dsp_value(input_var, region_data)

                    equation = equation.replace(input_var, str(value))

                # Evaluate the equation
                if "None" in equation:
                    soi_value = None
                else:
                    soi_value = eval(equation)

                    if soi_var_name.startswith("number_of"):
                        soi_value = round(soi_value)

            # cases when its directly a variable from DSP
            else:
                dsp_value = get_dsp_value(equation, region_data)
                soi_value = dsp_value

        # some of the ratio calculations have 0/(0+0). This should result in 0
        except ZeroDivisionError:
            soi_value = 0

        soi_df.loc[len(soi_df)] = {
            "soi_name": soi_name,
            "var_name": soi_var_name,
            "soi_description": soi_var_description,
            "methodology": methodology,
            "SECAP_link": SECAP_link,
            "SDG_targets": SDG_targets,
            "var_unit": var_unit,
            "value": soi_value,
        }

    # SOI calculations using other SOIs - Totals
    soi_vars_with_totals = soi_metadata_df[soi_metadata_df["data_source"] == "TOTAL"][
        [
            "soi_name",
            "var_name",
            "soi_description",
            "methodology",
            "SECAP_link",
            "SDG_targets",
            "var_unit",
            "calculation",
        ]
    ]

    for idx, row in soi_vars_with_totals.iterrows():

        soi_name = row["soi_name"]
        soi_var_name = row["var_name"]
        soi_var_description = row["soi_description"]
        methodology = row["methodology"]
        SECAP_link = row["SECAP_link"]
        SDG_targets = row["SDG_targets"]
        var_unit = row["var_unit"]
        equation = row["calculation"]

        equation = equation.replace("\n", " ")

        input_vars = extract_variables(equation)

        for input_var in input_vars:
            value = soi_df[soi_df["var_name"] == input_var]["value"].item()

            equation = equation.replace(input_var, str(value))

        value = eval(equation)
        soi_value = value

        soi_df.loc[len(soi_df)] = {
            "soi_name": soi_name,
            "var_name": soi_var_name,
            "soi_description": soi_var_description,
            "methodology": methodology,
            "SECAP_link": SECAP_link,
            "SDG_targets": SDG_targets,
            "var_unit": var_unit,
            "value": soi_value,
        }

    # save calculated SOIs as excel
    soi_df.to_excel(
        os.path.join(current_dir, "data", "output", f"SOIs_{region_code}.xlsx"),
        index=False,
    )
    return soi_df

def get_secap_filling_positions():
    """
    Get the filling positions for the CoM template
    
    use template COM file and variables_with_details_and_tags.xlsx to get the filling positions
    """

    # reading CoM template file
    original_file_path = os.path.join(
        current_dir, "data", "input", "CoM-Europe_reporting_template_2023_v4.xlsx"
    )

    # reading admin_business_and_social_KPIs sheet
    soi_metadata_df = pd.read_excel(
        os.path.join(
            current_dir,
            "data",
            "input",
            "variables_with_details_and_tags.xlsx",
        ),
            sheet_name="admin_business_and_social_KPIs",
        )
    logger.info(f"Total SOIs: {len(soi_metadata_df)}")
        # Open the copied file
    try:
        workbook = load_workbook(original_file_path)
    except Exception as e:
        logger.error(f"Failed to load workbook: {str(e)}")
        raise
    
    # fill sheets
    sheets_to_fill = [
            "GHG emissions",
            "Risks & vulnerabilities",
            "Energy poverty assessment",
        ]

    secap_filling_positions = {"GHG emissions": {}, "Risks & vulnerabilities": {}, "Energy poverty assessment": {}}
    for sheet_name in sheets_to_fill:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_column = min(sheet.max_column, 26)  # Limit to column Z
        logger.info(f"Reading sheet: {sheet_name}, max row: {max_row}, max column: {max_column}")
        for row in sheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_column
        ):
            for cell in row:
                if cell.value in soi_metadata_df["var_name"].values:
                    logger.info(f"row: {cell.row}, column: {cell.column}")
                    secap_filling_positions[sheet_name][cell.value] = {"position": [cell.row, cell.column]}

    input_dir = os.path.join(current_dir, "data", "input")
    with open(os.path.join(input_dir, f"secap_filling_positions_template.json"), "w") as f:
        logger.info(f"Saving secap filling positions to {os.path.join(input_dir, f'secap_filling_positions_template.json')}")
        json.dump(secap_filling_positions, f, indent=4)

def convert_soi_vars_excel_to_json():
    """
    Convert SOI variables from excel to json
    """
    soi_metadata_df = pd.read_excel(
        os.path.join(
            current_dir,
            "data",
            "input",
            "variables_with_details_and_tags.xlsx",
        ),
        sheet_name="admin_business_and_social_KPIs",
        usecols=["soi_name", "var_name"]
    )
    soi_json = json.loads(soi_metadata_df.to_json(orient="records"))
    soi_json_modified = {}
    for i in range(len(soi_json)):
        soi_json_modified[soi_json[i]["var_name"]] = soi_json[i]["soi_name"]

    with open(os.path.join(current_dir, "data", "input", f"soi_vars.json"), "w") as f:
        json.dump(soi_json_modified, f, indent=4)

def merge_soi_vars_json_with_secap_filling_positions():
    """
    Merge SOI variables from json with secap filling positions

    The main reason for doing this: there is no one-to-one mapping between the SOI variable names and the CoM template positions.
    This helps to show to the user in the right order as they are in the CoM template.
    """
    with open(os.path.join(current_dir, "data", "input", f"soi_vars.json"), "r") as f:
        soi_json_modified = json.load(f)

    with open(os.path.join(current_dir, "data", "input", f"secap_filling_positions_template.json"), "r") as f:
        secap_filling_positions = json.load(f)

    for sheet_name in secap_filling_positions.keys():
        logger.info(f"Sheet name: {sheet_name}")
        for soi_var in secap_filling_positions[sheet_name].keys():
            secap_filling_positions[sheet_name][soi_var]["name"] = soi_json_modified[soi_var]
    
    with open(os.path.join(current_dir, "data", "input", f"secap_filling_positions_template_with_soi_names.json"), "w") as f:
        json.dump(secap_filling_positions, f, indent=4)

def fill_actions_sheet(region_code="ES511_08019"):
    """
    Fill the actions sheet with the calculated values.
    """ 
    wb = load_workbook(os.path.join(current_dir, "data", "input", "CoM-Europe_reporting_template_2023_v4.xlsx"))
    actions_sheet = wb["Actions"]
    actions_sheet["C16"].value = "title of the action test"
    wb.save(os.path.join(current_dir, "data", "output", f"CoM_{region_code}_actions.xlsx"))

def get_active_dimensions(workbook_path):
    """
    Get the active dimensions of the CoM template
    """
    wb = load_workbook(workbook_path, data_only=True)
    logger.info(f"Workbook path: {workbook_path}")
    logger.info(f"Sheet names: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        used_range = wb[sheet_name].calculate_dimension()
        logger.info(f"Used range for {sheet_name}: {used_range}")

def clean_com_template(workbook_path):
    """
    Clean the CoM template
    """
    wb = load_workbook(workbook_path)


def fill_com_template(region_code, soi_df, region_data, sheet_name):
    """
    Fill the CoM template with calculated values.
    """
    logger.info(f"Starting CoM template filling for {region_code}")
    start = time.time() 

    try:
        # Define file paths
        original_file_path = os.path.join(
            current_dir, "data", "input", "CoM-Europe_reporting_template_2023_v4.xlsx"
        )
        logger.info(f"Original file path: {original_file_path}")

        # Open the copied file
        try:
            workbook = load_workbook(original_file_path)
            for sheet in workbook.worksheets:
                sheet.data_validations.dataValidation = []  # remove validations

            # Save to a new file (or overwrite if you're sure)
            workbook.save(os.path.join(current_dir, "data", "input", f"CoM_cleaned_v4.xlsx"))
        except Exception as e:
            logger.error(f"Failed to load workbook: {str(e)}")
            raise
        
        output_dir = os.path.join(current_dir, "data", "output")
        if sheet_name == "all_sheets":
            output_file_path = os.path.join(output_dir, f"CoM_{region_code}.xlsx")
        else:
            output_file_path = os.path.join(output_dir, f"CoM_{region_code}_{sheet_name}.xlsx")

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        cleaned_file_path = os.path.join(current_dir, "data", "input", f"CoM_cleaned_v4.xlsx")

        # Make a copy of the original file
        try:
            shutil.copy2(cleaned_file_path, output_file_path)
            logger.info(f"Template copied to {output_file_path}")
        except Exception as e:
            logger.error(f"Failed to copy template: {str(e)}")
            raise

        # Open the copied file
        try:
            workbook = load_workbook(output_file_path)
        except Exception as e:
            logger.error(f"Failed to load workbook: {str(e)}")
            raise
        
        # fill sheets
        if sheet_name == "all_sheets":
            sheets_to_fill = [
                "GHG emissions",
                "Risks & vulnerabilities",
                "Energy poverty assessment",
            ]
        else:
            sheets_to_fill = [sheet_name]

        # secap_filling_positions = {"GHG emissions": {}, "Risks & vulnerabilities": {}, "Energy poverty assessment": {}}
        input_dir = os.path.join(current_dir, "data", "input")
        try:
            with open(os.path.join(input_dir, f"secap_filling_positions_template.json"), "r") as f:
                secap_filling_positions = json.load(f)
        except Exception as e:
            logger.error(f"Failed to load secap filling positions: {str(e)}")
            raise

        for sheet_name in sheets_to_fill:
            try:
                sheet = workbook[sheet_name]
                
                logger.info(f"Filling sheet: {sheet_name}")
                # TODO: replace with JSON file positions
                # Get used range
                max_row = sheet.max_row
                logger.info(f"Max row: {max_row}")
                max_column = min(sheet.max_column, 26)  # Limit to column Z
                logger.info(f"Max column: {max_column}")
                n_items_filled = 0
                for row in sheet.iter_rows(
                    min_row=1, max_row=max_row, min_col=1, max_col=max_column
                ):
                    for cell in row:
                        if cell.value in soi_df["var_name"].values:
                            logger.info(f"row: {cell.row}, column: {cell.column}")
                            #secap_filling_positions[sheet_name][cell.value] = [cell.row, cell.column]
                            #logger.info(soi_df[soi_df["var_name"] == cell.value])
                            cell.value = soi_df[soi_df["var_name"] == cell.value][
                                "value"
                            ].item()

                            n_items_filled = n_items_filled + 1

                        elif (
                            isinstance(cell.value, str)
                            and cell.value in region_data["var_name"].values
                        ):
                            dsp_value = get_dsp_value(cell.value, region_data)
                            cell.value = dsp_value

                            n_items_filled = n_items_filled + 1

                        # elif cell.value == "title":
                        #     logger.info(f"Filling actions sheet", cell.row, cell.column, cell.coordinate)
                        #     # TODO: fill actions sheet
                        #     cell.value = "title of the action test"
                            
                logger.info(
                    f"Finished filling {sheet_name}, Number of items filled = {n_items_filled}"
                )
            except Exception as e:
                logger.error(f"Error filling sheet {sheet_name}: {str(e)}")
                raise

        # comment it out after an initial run
        # this is only to speed up things after the first run
        # with open(os.path.join(output_dir, f"secap_filling_positions.json"), "w") as f:
        #     json.dump(secap_filling_positions, f, indent=4)

        # Save and close the workbook
        try:
            # Try to save with a temporary name first
            temp_path = output_file_path + ".tmp"
            workbook.save(temp_path)

            # If save successful, rename to final name
            if os.path.exists(output_file_path):
                os.remove(output_file_path)
            os.rename(temp_path, output_file_path)

            logger.info(f"Successfully saved workbook to {output_file_path}")
        except Exception as e:
            logger.error(f"Failed to save workbook: {str(e)}")
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise
        finally:
            workbook.close()

        end = time.time()
        logger.info(f"Template filling completed in {end-start:.2f} seconds")

        return output_file_path

    except Exception as e:
        logger.error(f"Failed to fill template: {str(e)}")
        raise RuntimeError(f"Failed to fill CoM template: {str(e)}")


if __name__ == "__main__":
    # get_secap_filling_positions()
    # convert_soi_vars_excel_to_json()
    # merge_soi_vars_json_with_secap_filling_positions()
    # region_code = "ES511_08019"
    # region_data = get_region_data(region_code)
    # soi_df = calculate_sois(region_code, region_data)
    # fill_com_template(region_code, soi_df, region_data, sheet_name="all_sheets")
    #fill_actions_sheet(region_code="ES511_08019")
    get_active_dimensions(os.path.join(current_dir, "data", "input", "CoM_cleaned_v4.xlsx"))